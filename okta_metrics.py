import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta, timezone
from collections import Counter
import os

# ── CONFIG ──────────────────────────────────────────
OKTA_DOMAIN = "https://usac-admin.okta.com"   # <-- replace this
API_TOKEN   = ""             # <-- replace this
# ────────────────────────────────────────────────────

# ── DATE RANGE (April 1 - May 1, 2026) ──────────────
since = "2026-04-01T00:00:00Z"
until = "2026-05-01T23:59:59Z"

# ── FETCH EVENTS ─────────────────────────────────────
url = f"{OKTA_DOMAIN}/api/v1/logs"
params = {
    "filter": 'eventType eq "user.account.update_profile"',
    "since": since,
    "until": until,
    "limit": 1000
}
headers = {
    "Authorization": f"SSWS {API_TOKEN}",
    "Accept": "application/json"
}

print("Fetching events from Okta System Log...")
print(f"  Date range: {since} to {until}")

events = []
while url:
    resp = requests.get(url, headers=headers, params=params)

    if resp.status_code != 200:
        print(f"Error: {resp.status_code} - {resp.text}")
        break

    data = resp.json()
    events.extend(data)
    print(f"  Fetched {len(events)} events so far...")

    url = None
    params = {}
    for part in resp.headers.get("Link", "").split(","):
        if 'rel="next"' in part:
            url = part.split(";")[0].strip().strip("<>")
            break

print(f"\nDone. Total events collected: {len(events)}")

# ── CALCULATE METRICS + EXPORT ───────────────────────
if events:
    daily_counts = Counter()
    for event in events:
        date = event["published"][:10]
        daily_counts[date] += 1

    avg_per_day = sum(daily_counts.values()) / len(daily_counts)
    peak_day    = max(daily_counts, key=daily_counts.get)

    print("\n─────────────────────────────────────")
    print(f"  Date range        : Apr 1 – May 1, 2026")
    print(f"  Total updates     : {len(events)}")
    print(f"  Days with activity: {len(daily_counts)}")
    print(f"  Avg updates/day   : {avg_per_day:.1f}")
    print(f"  Peak day          : {peak_day}")
    print(f"  Peak day count    : {daily_counts[peak_day]}")
    print("─────────────────────────────────────")
    print("\n  Each update = 1 potential Event Hook trigger")

    # ── EXPORT TO EXCEL ──────────────────────────────
    output_path = os.path.join(os.path.dirname(__file__), "okta_profile_updates.xlsx")
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center")

    # ── TAB 1: SUMMARY ───────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    summary_data = [
        ["Metric", "Value"],
        ["Date Range", "Apr 1 – May 1, 2026"],
        ["Total Updates", len(events)],
        ["Days With Activity", len(daily_counts)],
        ["Avg Updates Per Day", f"{avg_per_day:.1f}"],
        ["Peak Day", peak_day],
        ["Peak Day Count", daily_counts[peak_day]],
        ["Est. Avg Hook Triggers/Hour", f"{avg_per_day / 24:.1f}"],
        ["Est. Peak Hook Triggers/Hour", f"{daily_counts[peak_day] / 24:.1f}"],
    ]

    for row in summary_data:
        ws1.append(row)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 25

    # ── TAB 2: DAILY BREAKDOWN ───────────────────────
    ws2 = wb.create_sheet("Daily Breakdown")
    ws2.append(["Date", "Profile Update Count"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for date in sorted(daily_counts.keys()):
        ws2.append([date, daily_counts[date]])
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 25

    # ── TAB 3: RAW EVENTS ────────────────────────────
    ws3 = wb.create_sheet("Raw Events")
    ws3.append(["Timestamp", "Event Type", "Actor", "Actor ID", "Target User", "Target User ID", "Client IP", "Outcome"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for event in events:
        targets     = event.get("target", [])
        target_name = targets[0].get("displayName", "") if targets else ""
        target_id   = targets[0].get("id", "") if targets else ""

        ws3.append([
            event.get("published", ""),
            event.get("eventType", ""),
            event.get("actor", {}).get("displayName", ""),
            event.get("actor", {}).get("id", ""),
            target_name,
            target_id,
            event.get("client", {}).get("ipAddress", ""),
            event.get("outcome", {}).get("result", ""),
        ])

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["D"].width = 25
    ws3.column_dimensions["E"].width = 25
    ws3.column_dimensions["F"].width = 25
    ws3.column_dimensions["G"].width = 20
    ws3.column_dimensions["H"].width = 15

    wb.save(output_path)
    print(f"\n  Excel file saved to: {output_path}")

else:
    print("No profile update events found in this date range.")